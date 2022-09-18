from ast import Lambda
from ntpath import join
from tkinter import N
import pandas as pd
import uuid
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import numpy as np
import os


def mper(val):
    return str(round(100*val))+'%'

#threshold of number of isolates tested for given antibiotic out of total number of isolates of this bacteria. If it is below the threshold we are not showing a result
thresh=0.02

list_of_versions = pd.read_csv(os.path.join('outdata','list_of_versions.csv'))

for _, ver in list_of_versions.iterrows():
    prov = ver['provider']
    loc = ver['location']

    prepo = pd.read_csv(os.path.join('outdata',f'pre-processed-{prov}-{loc}.csv'))
    prepo['pathogen'] = prepo['pathogen'].str.capitalize()
    noi = pd.read_csv(os.path.join('outdata',f'number-of-isolates-{prov}-{loc}.csv'))
    noi['pathogen'] = noi['pathogen'].str.capitalize()
    # rawsummary = pd.merge(prepo,noi, on = ['specimen_category','pathogen'])
    # easier to use to dfs for noi and rest of the data
    rawsummary = prepo[['specimen_category', 'pathogen', 'antibiotic', 'total', 'sensitivity']]

    #  ######
    #    ##
    #    ##
    #    ##
    #    ##
    #  ######
    #
    #  Calculate antibiogram
    #
    print("################################")
    print("################################")
    print(f"amb-{prov}-{loc}.xlsx")
    print("################################")
    print("################################")

    with pd.ExcelWriter(os.path.join("outdata",f"amb-{prov}-{loc}.xlsx"), engine="openpyxl") as writer:
        for specimen_type in prepo['specimen_category'].unique():
            print(specimen_type)
            df = rawsummary[rawsummary['specimen_category']==specimen_type]
            df_n = noi[noi['specimen_category']==specimen_type]
            df_n = df_n.sort_values(by=['number_of_isolates'], ascending = False)
            df_n = df_n.set_index('pathogen')

            amb_matrix_perc = pd.pivot(df, index='antibiotic', columns='pathogen', values='sensitivity')
            amb_matrix_tot = pd.pivot(df, index='antibiotic', columns='pathogen', values='total')

            new_row = dict()
            full_amb_list = []

            total_n_iso = df_n['number_of_isolates'].sum()
            new_row[('total','--')] = total_n_iso
            df_n = df_n.sort_values(by=['number_of_isolates'], ascending = False).head(5) # Taking only top 5 pathogens per specimen type


            for pat, niso in df_n.iterrows():
                new_row[(pat,'n')] = niso['number_of_isolates']
                new_row[(pat,'%')] = mper(niso['number_of_isolates'] / total_n_iso)
                
            
            new_row[('overall sensitivity','%')] = '--'
            full_amb_list.append(new_row)
            tots_row = new_row


            for (antibio, aperc), (_, atot) in zip(amb_matrix_perc.iterrows(), amb_matrix_tot.iterrows()):
                new_row = dict()
                new_row[('total','--')] = atot.sum()
                oversense = 0
                for pat, _ in df_n.iterrows():
                    if np.isnan(atot[pat]):
                        new_row[(pat,'n')] = 0
                        new_row[(pat,'%')] = '--'
                        
                    else:
                        #if number of isolates tested on this antibiotic i less than some percentage  of total number of isolates we will not report sensitivity
                        new_row[(pat,'n')] = atot[pat]
                        if (atot[pat] / tots_row[(pat,'n')]) < thresh:
                            new_row[(pat,'%')] = '--'
                        else:
                            new_row[(pat,'%')] = mper(aperc[pat])
                        n_times_per = atot[pat] * aperc[pat]
                        oversense = oversense + n_times_per
                new_row['overall sensitivity','%'] = mper(oversense/atot.sum())
                
                full_amb_list.append(new_row)
                

            full_amb_index = ['Total number of isolates'] + list(amb_matrix_perc.index)

            full_amb = pd.DataFrame.from_records(full_amb_list,index=full_amb_index)

            new_columns = pd.MultiIndex.from_tuples(full_amb.columns, names=['Pathogen', 'number of isolates and sensitivity'])
            full_amb.columns = new_columns
            
            ######delete later
            datacheck = 'check'
            ## Checks
            full_amb.to_csv(f'{datacheck}/fullamb.csv')
            #########
                
            # I want to create a separate database for "overall sensitivity" from the existing one
            # I want to remove all columns other than the name of antibiotcs and the overall sensitivity
            # and sort the database in descending order of overall sensitivity
            overall_sens = full_amb
            xnew_columns = pd.MultiIndex.from_tuples(overall_sens.columns, names=['Antibiotic', 'number of isolates and sensitivity'])
            overall_sens.columns = xnew_columns
   
            overall_sens = overall_sens.loc[:, ["overall sensitivity"]].sort_values(by=[("overall sensitivity", "%")], ascending=False).drop('Total number of isolates')
            
            full_amb.to_excel(writer,sheet_name=specimen_type, startrow=0, startcol=0)
            
            overall_sens.to_excel(writer,sheet_name=specimen_type, startrow=1, startcol=full_amb.shape[0]+1)
            #the line above still needs tweaking, in case of some xlsx files this second table gets pushed 9 or more columns to the right
            # Set backgrund colors depending on cell values
            worksheet = writer.sheets[specimen_type]


    #  ######   ######  ######
    #    ##       ##      ##
    #    ##       ##      ##
    #    ##       ##      ##
    #    ##       ##      ##
    #  ######   ######  ######
    #
    #  formatting
    #
            heading_size = 3 #I'm not setting it up anywhere it is for some reason longer than I'd want

            worksheet.row_dimensions[1].height = 150

            worksheet.column_dimensions[worksheet.cell(row=3,column=1).column_letter].width = 40
            for ccc in range(2,len(full_amb.columns)+2):
                if ccc%2:
                    worksheet.column_dimensions[worksheet.cell(row=3,column=ccc).column_letter].width = 5
                else:
                    worksheet.column_dimensions[worksheet.cell(row=3,column=ccc).column_letter].width = 5
                worksheet.cell(row=1,column=ccc).alignment = Alignment(textRotation=70)
                for rrr in range(heading_size+1,heading_size+1+len(full_amb)):
                    worksheet.cell(row=rrr,column=ccc).alignment = Alignment(vertical='center')
                    worksheet.cell(row=rrr,column=ccc).alignment = Alignment(horizontal='center')

            # worksheet.column_dimensions[worksheet.cell(row=3,column=len(full_amb.columns)+1).column_letter].width = 20
            # worksheet.column_dimensions[worksheet.cell(row=3,column=len(full_amb.columns)).column_letter].width = 20

            #starting after first entry which is pathogen wide totals
            for rrr in range(heading_size+2,heading_size+1+len(full_amb)):
                for ccc in list(range(4,len(full_amb.columns)+1,2))+[len(full_amb.columns)+1]:
                    value_str = worksheet.cell(row=rrr,column=ccc).value
                    if value_str == '--':
                        worksheet.cell(row=rrr,column=ccc).fill = PatternFill("solid", start_color=('ededed'))
                    else:
                        value = int(value_str.rstrip('%'))
                        worksheet.cell(row=rrr,column=ccc).fill = PatternFill("solid", start_color=('d62d20' if value < 50 else 'ed6f00' if value < 75 else 'eee600' if value < 90 else 'a0e040'))

            for ccc in range(1,len(full_amb.columns)+2):
                worksheet.cell(row=4,column=ccc).fill = PatternFill("solid", start_color='6495ed')
            worksheet.delete_rows(3,1)
            worksheet.cell(row=2, column=1).value=None # getting rid of the line "number of isolates and sensitivity"
            worksheet.cell(row=1,column=1).alignment = Alignment(horizontal='center', vertical='center') #fixing the alignment for "Pathogen"


############for overal sensitivity
            worksheet.column_dimensions[worksheet.cell(row=3,column=full_amb.shape[0]+2).column_letter].width = 40
            for ccc in range(2,len(overall_sens.columns)):
                if ccc%2:
                    worksheet.column_dimensions[worksheet.cell(row=3,column=ccc).column_letter].width = 5
                else:
                    worksheet.column_dimensions[worksheet.cell(row=3,column=ccc).column_letter].width = 5
                worksheet.cell(row=3,column=ccc).alignment = Alignment(horizontal='center', vertical='center')
                for rrr in range(heading_size+1,heading_size+1+len(overall_sens)):
                    worksheet.cell(row=rrr,column=ccc).alignment = Alignment(vertical='center')
                    worksheet.cell(row=rrr,column=ccc).alignment = Alignment(horizontal='center')

            


            #for rrr in range(heading_size+2,heading_size+1+len(overall_sens)):
             #   for ccc in list(range(2,len(overall_sens.columns)+1,2))+[len(overall_sens.columns)+1]:
              #      value_str = worksheet.cell(row=rrr,column=ccc).value
               #     if value_str == '--':
                #        worksheet.cell(row=rrr,column=ccc).fill = PatternFill("solid", start_color=('ededed'))
                 #   else:
                  #      value = int(value_str.rstrip('%'))
                   #     worksheet.cell(row=rrr,column=ccc).fill = PatternFill("solid", start_color=('d62d20' if value < 50 else 'ed6f00' if value < 75 else 'eee600' if value < 90 else 'a0e040'))